from flask import Flask,redirect,render_template,request,json,jsonify,session,url_for,flash,Markup
from flask_mail import Mail,Message
import mysql.connector
import requests
import random
import datetime
from datetime import datetime
from datetime import timedelta
from datetime import date
import hashlib
import uuid
import os
from werkzeug.utils import secure_filename
from bs4 import BeautifulSoup
# import cron
from flask_socketio import SocketIO,send,emit,join_room,leave_room
import yagmail
import xlrd
import razorpay
import xlsxwriter 

import convertapi

from publitio import PublitioAPI
import json

from openpyxl import load_workbook

publitio_api = PublitioAPI(key='1DJB25jRHW9WLQMvLlap', secret='asD7u807m6by7WZ36wStXJmowPqlzisb')

app = Flask(__name__)
app.secret_key = "1509"

mail=Mail(app)

app_base_url = "http://127.0.0.1:5000/"


UPLOAD_FOLDER = '/uploads/'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAIL_SERVER']='smtp.gmail.com'
app.config['MAIL_PORT'] = 465
app.config['MAIL_USERNAME'] = 'alumni.association.goa@gmail.com'
app.config['MAIL_PASSWORD'] = 'ats12345'
app.config['MAIL_USE_TLS'] = False
app.config['MAIL_USE_SSL'] = True

mail = Mail(app)


@app.route('/')
def index():
    cnx = mysql.connector.connect(host="localhost",user="root",passwd="",database="ats")
    con = cnx.cursor()
    
    sql = "SELECT * FROM cron_resume_extraction WHERE status='PENDING'"
    con.execute(sql)
    res = con.fetchall()
    try :
        cron_jon_id = int(res[0][0])
        # print(res)
            
        KEYWORDS = str(res[0][3]).split(",")
        
            
        file_loc = res[0][4]
        KEYWORDS = str(res[0][3]).split(",")
    except :
        return "<div style='position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);'><h1>No Record Found For Resume Extraction.</h1></div>"
    
    file = xlrd.open_workbook(file_loc)
    sheet = file.sheet_by_index(0)
    PATH = []             
    main_content = []
    k = 0
    for i in range(1,sheet.nrows):   
        p = sheet.cell_value(i,4)
        p = p.replace("http://127.0.0.1:5000",".")
        PATH.append(p)
    # file:///D:/ALUMNI_TRACKING_SYSTEM/static/uploads/alumni_job_resumes/3_ff7540c1_d3f6_11ea_9bfd_80000bda9a1b_DARSHs%20Resume%20(3).pdf

    print(PATH)  

    c_dirName = "static/uploads/alumni_extracted_resumes/"

    try :
        os.mkdir(c_dirName)
                        
    except FileExistsError:
        print("Directory " + c_dirName + " already exists")

    
    CONVERTED_PATH = []

    for i in range(len(PATH)) :
        # path = str(PATH[i]).replace("http://127.0.0.1:5000/","")
        # with open(path, 'rb') as f:
        #     x = publitio_api.create_file(file=f,
        #                      title='My title',
        #                      description='My description')
    
    # t = json.load
    
        # if len(x['url_preview']) > 0 :
        #     url_p = x['url_preview']
        #     print(url_p)
        con_path =str(PATH[i])+".txt"
        convertapi.api_secret = 'jOQbkzhJV2zFsPq2'
        convertapi.convert('txt', {
            'File': PATH[i]
        }, from_format = 'pdf').save_files(con_path)
        CONVERTED_PATH.append(con_path)
        # else :
        #     CONVERTED_PATH.append("N/A")
     
    wb = load_workbook(file_loc)
    
    sheet = wb.active 
    g1 = sheet.cell(row = 1, column = 7) 
    g1.value = "Match (%)"
    h1 = sheet.cell(row = 1, column = 8) 
    h1.value = "Keywords Found"       


    # KEYWORDS = ['python','programming','google']
    r = 2

    for i in range(len(CONVERTED_PATH)) :
        
        FOUND_KEYWORDS = []
        if CONVERTED_PATH[i] != "N/A" :
            
            print("="*50)
            print("File : ",CONVERTED_PATH[i])
                        
            for j in range(len(KEYWORDS)) :
                f = open(CONVERTED_PATH[i],'r',encoding='utf-8').read().find(KEYWORDS[j])
                        
                if f != -1 and KEYWORDS[j] not in FOUND_KEYWORDS:
                    print("Keyword : ",KEYWORDS[j])
                    FOUND_KEYWORDS.append(KEYWORDS[j])
            print(FOUND_KEYWORDS)
            if len(FOUND_KEYWORDS) == 0 :
                print("Match : "+"0 %")
                g1 = sheet.cell(row = r, column = 7) 
                g1.value = "0 %"
                h1 = sheet.cell(row = r, column = 8) 
                h1.value = str("-")

                r += 1
            else :
                x = len(KEYWORDS)
                y = len(FOUND_KEYWORDS)
                        
                percent = (100*y) / x
                percent = "{:.2f}".format(percent)

                print("Match : "+str(percent)+" %")
                
                g1 = sheet.cell(row = r, column = 7) 
                g1.value = str(percent)
                
                
                sk_found = ','.join([str(elem) for elem in FOUND_KEYWORDS]) 
                
                h1 = sheet.cell(row = r, column = 8) 
                h1.value = str(sk_found)
    
                r += 1
    
    wb.save(file_loc) 
    
    
    msg = Message('EXTRACTED DATA FROM JOB RESUMES', sender = ('Alumni Association Of Goa','alumni.association.goa@gmail.com'), recipients = ["pateldarsh422@gmail.com"])
                    
    msg.html = '''
                                                              <!DOCTYPE html PUBLIC "-//W3C//DTD XHTML 1.0 Transitional //EN" "http://www.w3.org/TR/xhtml1/DTD/xhtml1-transitional.dtd">

<html xmlns="http://www.w3.org/1999/xhtml" xmlns:o="urn:schemas-microsoft-com:office:office" xmlns:v="urn:schemas-microsoft-com:vml">

<head>
    <!--[if gte mso 9]><xml><o:OfficeDocumentSettings><o:AllowPNG/><o:PixelsPerInch>96</o:PixelsPerInch></o:OfficeDocumentSettings></xml><![endif]-->
    <meta content="text/html; charset=utf-8" http-equiv="Content-Type" />
    <meta content="width=device-width" name="viewport" />
    <!--[if !mso]><!-->
    <meta content="IE=edge" http-equiv="X-UA-Compatible" />
    <!--<![endif]-->
    <title></title>
    <!--[if !mso]><!-->
    <link href="https://fonts.googleapis.com/css?family=Lato" rel="stylesheet" type="text/css" />
    <!--<![endif]-->
    <style type="text/css">
        body {
            margin: 0;
            padding: 0;
        }
        
        table,
        td,
        tr {
            vertical-align: top;
            border-collapse: collapse;
        }
        
        * {
            line-height: inherit;
        }
        
        a[x-apple-data-detectors=true] {
            color: inherit !important;
            text-decoration: none !important;
        }
    </style>
    <style id="media-query" type="text/css">
        @media (max-width: 670px) {
            .block-grid,
            .col {
                min-width: 320px !important;
                max-width: 100% !important;
                display: block !important;
            }
            .block-grid {
                width: 100% !important;
            }
            .col {
                width: 100% !important;
            }
            .col>div {
                margin: 0 auto;
            }
            img.fullwidth,
            img.fullwidthOnMobile {
                max-width: 100% !important;
            }
            .no-stack .col {
                min-width: 0 !important;
                display: table-cell !important;
            }
            .no-stack.two-up .col {
                width: 50% !important;
            }
            .no-stack .col.num4 {
                width: 33% !important;
            }
            .no-stack .col.num8 {
                width: 66% !important;
            }
            .no-stack .col.num4 {
                width: 33% !important;
            }
            .no-stack .col.num3 {
                width: 25% !important;
            }
            .no-stack .col.num6 {
                width: 50% !important;
            }
            .no-stack .col.num9 {
                width: 75% !important;
            }
            .video-block {
                max-width: none !important;
            }
            .mobile_hide {
                min-height: 0px;
                max-height: 0px;
                max-width: 0px;
                display: none;
                overflow: hidden;
                font-size: 0px;
            }
            .desktop_hide {
                display: block !important;
                max-height: none !important;
            }
        }
    </style>
</head>

<body class="clean-body" style="margin: 0; padding: 0; -webkit-text-size-adjust: 100%; background-color: #F5F5F5;">
    <!--[if IE]><div class="ie-browser"><![endif]-->
    <table bgcolor="#F5F5F5" cellpadding="0" cellspacing="0" class="nl-container" role="presentation" style="table-layout: fixed; vertical-align: top; min-width: 320px; Margin: 0 auto; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; background-color: #F5F5F5; width: 100%;" valign="top" width="100%">
        <tbody>
            <tr style="vertical-align: top;" valign="top">
                <td style="word-break: break-word; vertical-align: top;" valign="top">
                    <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td align="center" style="background-color:#F5F5F5"><![endif]-->
                    <div style="background-color:transparent;">
                        <div class="block-grid" style="Margin: 0 auto; min-width: 320px; max-width: 650px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: transparent;">
                            <div style="border-collapse: collapse;display: table;width: 100%;background-color:transparent;">
                                <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:transparent;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:650px"><tr class="layout-full-width" style="background-color:transparent"><![endif]-->
                                <!--[if (mso)|(IE)]><td align="center" width="650" style="background-color:transparent;width:650px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:5px; padding-bottom:5px;"><![endif]-->
                                <div class="col num12" style="min-width: 320px; max-width: 650px; display: table-cell; vertical-align: top; width: 650px;">
                                    <div style="width:100% !important;">
                                        <!--[if (!mso)&(!IE)]><!-->
                                        <div style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:5px; padding-bottom:5px; padding-right: 0px; padding-left: 0px;">
                                            <!--<![endif]-->
                                            <table border="0" cellpadding="0" cellspacing="0" class="divider" role="presentation" style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;" valign="top" width="100%">
                                                <tbody>
                                                    <tr style="vertical-align: top;" valign="top">
                                                        <td class="divider_inner" style="word-break: break-word; vertical-align: top; min-width: 100%; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%; padding-top: 10px; padding-right: 10px; padding-bottom: 10px; padding-left: 10px;" valign="top">
                                                            <table align="center" border="0" cellpadding="0" cellspacing="0" class="divider_content" height="10" role="presentation" style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; border-top: 0px solid transparent; height: 10px; width: 100%;" valign="top" width="100%">
                                                                <tbody>
                                                                    <tr style="vertical-align: top;" valign="top">
                                                                        <td height="10" style="word-break: break-word; vertical-align: top; -ms-text-size-adjust: 100%; -webkit-text-size-adjust: 100%;" valign="top"><span></span></td>
                                                                    </tr>
                                                                </tbody>
                                                            </table>
                                                        </td>
                                                    </tr>
                                                </tbody>
                                            </table>
                                            <!--[if (!mso)&(!IE)]><!-->
                                        </div>
                                        <!--<![endif]-->
                                    </div>
                                </div>
                                <!--[if (mso)|(IE)]></td></tr></table><![endif]-->
                                <!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
                            </div>
                        </div>
                    </div>
                    <div style="background-color:transparent;">
                        <div class="block-grid two-up no-stack" style="Margin: 0 auto; min-width: 320px; max-width: 650px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: #FFFFFF;">
                            <div style="border-collapse: collapse;display: table;width: 100%;background-color:#FFFFFF;">
                                <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:transparent;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:650px"><tr class="layout-full-width" style="background-color:#FFFFFF"><![endif]-->
                                <!--[if (mso)|(IE)]><td align="center" width="325" style="background-color:#FFFFFF;width:325px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 25px; padding-top:25px; padding-bottom:25px;"><![endif]-->
                                <div class="col num6" style="min-width: 320px; max-width: 325px; display: table-cell; vertical-align: top; width: 325px;">
                                    <div style="width:100% !important;">
                                        <!--[if (!mso)&(!IE)]><!-->
                                        <div style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:25px; padding-bottom:25px; padding-right: 0px; padding-left: 25px;">
                                            <!--<![endif]-->
                                            <div align="left" class="img-container left fixedwidth" style="padding-right: 0px;padding-left: 0px;">
                                                <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="padding-right: 0px;padding-left: 0px;" align="left"><![endif]--><img alt="Image" border="0" class="left fixedwidth" src="http://cdn.mcauto-images-production.sendgrid.net/92f4cab4130b7228/6ac70d4b-6462-4a67-ac30-faec499c6d4f/256x105.png" style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 195px; display: block;" title="Image" width="195" />
                                                <!--[if mso]></td></tr></table><![endif]-->
                                            </div>
                                            <!--[if (!mso)&(!IE)]><!-->
                                        </div>
                                        <!--<![endif]-->
                                    </div>
                                </div>
                                <!--[if (mso)|(IE)]></td></tr></table><![endif]-->
                                <!--[if (mso)|(IE)]></td><td align="center" width="325" style="background-color:#FFFFFF;width:325px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 25px; padding-left: 0px; padding-top:25px; padding-bottom:25px;"><![endif]-->
                                <div class="col num6" style="min-width: 320px; max-width: 325px; display: table-cell; vertical-align: top; width: 325px;">
                                    <div style="width:100% !important;">
                                        <!--[if (!mso)&(!IE)]><!-->
                                        <div style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:25px; padding-bottom:25px; padding-right: 25px; padding-left: 0px;">
                                            <!--<![endif]-->
                                            <div align="right" class="button-container" style="padding-top:10px;padding-right:0px;padding-bottom:10px;padding-left:10px;">
                                                <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-spacing: 0; border-collapse: collapse; mso-table-lspace:0pt; mso-table-rspace:0pt;"><tr><td style="padding-top: 10px; padding-right: 0px; padding-bottom: 10px; padding-left: 10px" align="right"><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="#" style="height:31.5pt; width:115.5pt; v-text-anchor:middle;" arcsize="34%" stroke="false" fillcolor="#0088ff"><w:anchorlock/><v:textbox inset="0,0,0,0"><center style="color:#ffffff; font-family:Tahoma, Verdana, sans-serif; font-size:14px"><![endif]-->
                                               
                                                <!--[if mso]></center></v:textbox></v:roundrect></td></tr></table><![endif]-->
                                            </div>
                                            <!--[if (!mso)&(!IE)]><!-->
                                        </div>
                                        <!--<![endif]-->
                                    </div>
                                </div>
                                <!--[if (mso)|(IE)]></td></tr></table><![endif]-->
                                <!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
                            </div>
                        </div>
                    </div>
                    <div style="background-color:transparent;">
                        <div class="block-grid" style="Margin: 0 auto; min-width: 320px; max-width: 650px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: #D6E7F0;">
                            <div style="border-collapse: collapse;display: table;width: 100%;background-color:#D6E7F0;">
                                <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:transparent;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:650px"><tr class="layout-full-width" style="background-color:#D6E7F0"><![endif]-->
                                <!--[if (mso)|(IE)]><td align="center" width="650" style="background-color:#D6E7F0;width:650px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 25px; padding-left: 25px; padding-top:5px; padding-bottom:60px;"><![endif]-->
                                <div class="col num12" style="min-width: 320px; max-width: 650px; display: table-cell; vertical-align: top; width: 650px;">
                                    <div style="width:100% !important;">
                                        <!--[if (!mso)&(!IE)]><!-->
                                        <div style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:5px; padding-bottom:60px; padding-right: 25px; padding-left: 25px;">
                                            <!--<![endif]-->
                                            <div align="center" class="img-container center fixedwidth" style="padding-right: 0px;padding-left: 0px;">
                                                <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr style="line-height:0px"><td style="padding-right: 0px;padding-left: 0px;" align="center"><![endif]-->
                                                <div style="font-size:1px;line-height:45px"> </div><img align="center" alt="Image" border="0" class="center fixedwidth" src="http://cdn.mcauto-images-production.sendgrid.net/92f4cab4130b7228/48bc327b-a5f2-40e0-80d4-ff98d66ffec1/1129x619.png" style="text-decoration: none; -ms-interpolation-mode: bicubic; border: 0; height: auto; width: 100%; max-width: 540px; display: block;" title="Image" width="540" />
                                                <!--[if mso]></td></tr></table><![endif]-->
                                            </div>
                                            <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 10px; padding-left: 15px; padding-top: 20px; padding-bottom: 0px; font-family: Tahoma, Verdana, sans-serif"><![endif]-->
                                            <div style="color:#052d3d;font-family:Lato, Tahoma, Verdana, Segoe, sans-serif;line-height:1.5;padding-top:20px;padding-right:10px;padding-bottom:0px;padding-left:15px;">
                                                <div style="font-size: 12px; line-height: 1.5; font-family: Lato, Tahoma, Verdana, Segoe, sans-serif; color: #052d3d; mso-line-height-alt: 18px;">
                                                  
                                                </div>
                                            </div>
                                            <!--[if mso]></td></tr></table><![endif]-->
                                            <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 10px; padding-left: 10px; padding-top: 10px; padding-bottom: 10px; font-family: Tahoma, Verdana, sans-serif"><![endif]-->
                                            <div style="color:#555555;font-family:Lato, Tahoma, Verdana, Segoe, sans-serif;line-height:1.2;padding-top:10px;padding-right:10px;padding-bottom:10px;padding-left:10px;">
                                                <div style="font-size: 12px; line-height: 1.2; color: #555555; font-family: Lato, Tahoma, Verdana, Segoe, sans-serif; mso-line-height-alt: 14px;">
                                                    <p style="font-size: 18px; line-height: 1.2; text-align: center; word-break: break-word; mso-line-height-alt: 22px; margin: 0;"><span style="font-size: 18px; color: #000000;">Your Extracted Data Is Now Ready. Click On Download Button To Download.</span></p>
                                                </div>
                                            </div>
                                            <!--[if mso]></td></tr></table><![endif]-->
                                            <div align="center" class="button-container" style="padding-top:20px;padding-right:10px;padding-bottom:10px;padding-left:10px;">
                                                <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-spacing: 0; border-collapse: collapse; mso-table-lspace:0pt; mso-table-rspace:0pt;"><tr><td style="padding-top: 20px; padding-right: 10px; padding-bottom: 10px; padding-left: 10px" align="center"><v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="#" style="height:39pt; width:220.5pt; v-text-anchor:middle;" arcsize="29%" stroke="false" fillcolor="#fc7318"><w:anchorlock/><v:textbox inset="0,0,0,0"><center style="color:#ffffff; font-family:Tahoma, Verdana, sans-serif; font-size:16px"><![endif]--><a href="'''+app_base_url+str(file_loc)+'''" style="-webkit-text-size-adjust: none; text-decoration: none; display: inline-block; color: #ffffff; background-color: #fc7318; border-radius: 15px; -webkit-border-radius: 15px; -moz-border-radius: 15px; width: auto; width: auto; border-top: 1px solid #fc7318; border-right: 1px solid #fc7318; border-bottom: 1px solid #fc7318; border-left: 1px solid #fc7318; padding-top: 10px; padding-bottom: 10px; font-family: Lato, Tahoma, Verdana, Segoe, sans-serif; text-align: center; mso-border-alt: none; word-break: keep-all;" target="_blank"><span style="padding-left:40px;padding-right:40px;font-size:16px;display:inline-block;"><span style="font-size: 16px; line-height: 2; word-break: break-word; mso-line-height-alt: 32px;"><strong>DOWNLOAD NOW</strong></span></span></a>
                                                <!--[if mso]></center></v:textbox></v:roundrect></td></tr></table><![endif]-->
                                            </div>
                                            <!--[if (!mso)&(!IE)]><!-->
                                        </div>
                                        <!--<![endif]-->
                                    </div>
                                </div>
                                <!--[if (mso)|(IE)]></td></tr></table><![endif]-->
                                <!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
                            </div>
                        </div>
                    </div>
                    <div style="background-color:transparent;">
                        <div class="block-grid" style="Margin: 0 auto; min-width: 320px; max-width: 650px; overflow-wrap: break-word; word-wrap: break-word; word-break: break-word; background-color: transparent;">
                            <div style="border-collapse: collapse;display: table;width: 100%;background-color:transparent;">
                                <!--[if (mso)|(IE)]><table width="100%" cellpadding="0" cellspacing="0" border="0" style="background-color:transparent;"><tr><td align="center"><table cellpadding="0" cellspacing="0" border="0" style="width:650px"><tr class="layout-full-width" style="background-color:transparent"><![endif]-->
                                <!--[if (mso)|(IE)]><td align="center" width="650" style="background-color:transparent;width:650px; border-top: 0px solid transparent; border-left: 0px solid transparent; border-bottom: 0px solid transparent; border-right: 0px solid transparent;" valign="top"><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 0px; padding-left: 0px; padding-top:20px; padding-bottom:60px;"><![endif]-->
                                <div class="col num12" style="min-width: 320px; max-width: 650px; display: table-cell; vertical-align: top; width: 650px;">
                                    <div style="width:100% !important;">
                                        <!--[if (!mso)&(!IE)]><!-->
                                        <div style="border-top:0px solid transparent; border-left:0px solid transparent; border-bottom:0px solid transparent; border-right:0px solid transparent; padding-top:20px; padding-bottom:60px; padding-right: 0px; padding-left: 0px;">
                                            <!--<![endif]-->
                                            <table cellpadding="0" cellspacing="0" class="social_icons" role="presentation" style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt;" valign="top" width="100%">
                                                <tbody>
                                                    <tr style="vertical-align: top;" valign="top">
                                                        <td style="word-break: break-word; vertical-align: top; padding-top: 10px; padding-right: 10px; padding-bottom: 10px; padding-left: 10px;" valign="top">
                                                            <table align="center" cellpadding="0" cellspacing="0" class="social_table" role="presentation" style="table-layout: fixed; vertical-align: top; border-spacing: 0; border-collapse: collapse; mso-table-tspace: 0; mso-table-rspace: 0; mso-table-bspace: 0; mso-table-lspace: 0;" valign="top">
                                                                <tbody>
                                                                    <tr align="center" style="vertical-align: top; display: inline-block; text-align: center;" valign="top">
                                                                        <td style="word-break: break-word; vertical-align: top; padding-bottom: 5px; padding-right: 8px; padding-left: 8px;" valign="top">
                                                                            <a href="https://www.facebook.com/darshpatelofficial" target="_blank"><img alt="Facebook" height="32" src="http://cdn.mcauto-images-production.sendgrid.net/92f4cab4130b7228/93ef56de-a14b-43c0-8cda-7ba7f3b69b19/64x64.png" style="text-decoration: none; -ms-interpolation-mode: bicubic; height: auto; border: none; display: block;" title="Facebook" width="32" /></a>
                                                                        </td>
                                                                        <td style="word-break: break-word; vertical-align: top; padding-bottom: 5px; padding-right: 8px; padding-left: 8px;" valign="top">
                                                                            <a href="https://twitter.com/pateldarsh10" target="_blank"><img alt="Twitter" height="32" src="http://cdn.mcauto-images-production.sendgrid.net/92f4cab4130b7228/81c04100-cf31-4016-8820-9bf9296d39a0/64x64.png" style="text-decoration: none; -ms-interpolation-mode: bicubic; height: auto; border: none; display: block;" title="Twitter" width="32" /></a>
                                                                        </td>
                                                                        <td style="word-break: break-word; vertical-align: top; padding-bottom: 5px; padding-right: 8px; padding-left: 8px;" valign="top">
                                                                            <a href="https://instagram.com/pateldarsh_007" target="_blank"><img alt="Instagram" height="32" src="http://cdn.mcauto-images-production.sendgrid.net/92f4cab4130b7228/858aeae3-142f-4b84-8400-87cc7e0e92d8/64x64.png" style="text-decoration: none; -ms-interpolation-mode: bicubic; height: auto; border: none; display: block;" title="Instagram" width="32" /></a>
                                                                        </td>
                                                                        <td style="word-break: break-word; vertical-align: top; padding-bottom: 5px; padding-right: 8px; padding-left: 8px;" valign="top">
                                                                            <a href="mailto:mailto:alumni.association.goa@gmail.com" target="_blank"><img alt="E-Mail" height="32" src="http://cdn.mcauto-images-production.sendgrid.net/92f4cab4130b7228/89016f33-3bcf-4dc9-b8ee-a4b5bc249ea4/64x64.png" style="text-decoration: none; -ms-interpolation-mode: bicubic; height: auto; border: none; display: block;" title="E-Mail" width="32" /></a>
                                                                        </td>
                                                                    </tr>
                                                                </tbody>
                                                            </table>
                                                        </td>
                                                    </tr>
                                                </tbody>
                                            </table>
                                            <!--[if mso]><table width="100%" cellpadding="0" cellspacing="0" border="0"><tr><td style="padding-right: 10px; padding-left: 10px; padding-top: 10px; padding-bottom: 10px; font-family: Tahoma, Verdana, sans-serif"><![endif]-->
                                            <div style="color:#555555;font-family:Lato, Tahoma, Verdana, Segoe, sans-serif;line-height:1.5;padding-top:10px;padding-right:10px;padding-bottom:10px;padding-left:10px;">
                                                <div style="font-size: 12px; line-height: 1.5; color: #555555; font-family: Lato, Tahoma, Verdana, Segoe, sans-serif; mso-line-height-alt: 18px;">
                                                    <p style="font-size: 14px; line-height: 1.5; text-align: center; word-break: break-word; mso-line-height-alt: 21px; margin: 0;">© 2020 - Alumni Association Of Goa - All Rights Reserved.</p>
                                                </div>
                                            </div>
                                            <!--[if mso]></td></tr></table><![endif]-->
                                            <!--[if (!mso)&(!IE)]><!-->
                                        </div>
                                        <!--<![endif]-->
                                    </div>
                                </div>
                                <!--[if (mso)|(IE)]></td></tr></table><![endif]-->
                                <!--[if (mso)|(IE)]></td></tr></table></td></tr></table><![endif]-->
                            </div>
                        </div>
                    </div>
                    <!--[if (mso)|(IE)]></td></tr></table><![endif]-->
                </td>
            </tr>
        </tbody>
    </table>
    <!--[if (IE)]></div><![endif]-->
</body>

</html>
                '''
    try :
        mail.send(msg)
        print("Mail Sent")
        
        sql = "DELETE FROM cron_resume_extraction WHERE id=%d" % (cron_jon_id)
        con.execute(sql)
        cnx.commit()
        print("CRON DELETED.")
    except :
        print("Error In Sending Mail.")
    


    # for i in res :
    # row = 1
    # records = 1
    # for i in res :
    #     column = 0
    #     worksheet.write(row, column, str(records)) 
    #     column += 1
    #     worksheet.write(row, column, str(i[3])) 
    #     column += 1
    #     worksheet.write(row, column, str(i[4]))
    #     column += 1
    #     worksheet.write(row, column, str("+")+str(i[5])) 
    #     column += 1
    #     worksheet.write(row, column, str("http://127.0.0.1:5000/")+str(i[6]))
    #     column += 1
    #     t = i[7].strftime("%Y-%m-%d %I:%M:%S %p")
    #     worksheet.write(row, column, str(t))
    #     row += 1
    #     records += 1
    # wb.close() 

                        

        
    con.close()
    return "<div style='position:absolute;top:50%;left:50%;transform:translate(-50%,-50%);'><h1>CRON Job Is Executing For Resume Extraction.</h1></div>"
    
if __name__ == "__main__" :
    app.run(debug=True,port=5002)